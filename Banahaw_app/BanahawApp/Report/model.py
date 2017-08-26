import datetime
import operator
from openpyxl import Workbook, drawing
from openpyxl.styles import Color, PatternFill, Font, Alignment
from openpyxl.cell import Cell
import os
import re
from flask import request
from sqlalchemy import and_
from BanahawApp import Session,Mini_func, app
from BanahawApp.table import T_Transaction, T_Member00, T_Attendants01, T_Attendants, T_Member01


class Attendant_report(object):

	def __init__(self, **kwargs):
		self.__session = Session()
		self.__args = kwargs
		self.__retval = None
		self.__tempdict = dict()
		self.__total = 0

	def get_reports_data(self):
		"""."""
		ds = self.__args.get('from',None)
		de = self.__args.get('to',None)
		by_id = self.__args.get('attendantid', None)

		if all([ds,de,by_id]):
			self.__tempdict['data'] = list()
			transactionstotal, transsales = self.__get_transactions(ds, de, by_id)
			memberstotal, memsales = self.__get_members(ds, de, by_id)
			upgradestotal, upgradessales = self.__get_members_upg(ds, de, by_id)

		if self.__tempdict:
			self.__retval = self.__tempdict
			self.__retval['total_on_service'] = transactionstotal
			self.__retval['total_on_membership'] = memberstotal + upgradestotal
			self.__retval['total_sales'] = transsales + memsales + upgradessales

		return self.__retval

	def __get_transactions(self, ds, de, attid):
		search_filter = list()

		search_filter.append(getattr(T_Transaction, 'datecreated').between(ds,de))
		search_filter.append(getattr(T_Transaction, 'attendantid') == attid)
		search_filter.append(getattr(T_Transaction, 'active') == 0)

		result = self.__session.query(T_Transaction).filter(
			and_(*search_filter)).all()

		total = 0
		total2 = 0

		for data in result:
			key = data.datecreated.strftime('%B-%d-%Y')

			# if key not in self.__tempdict:
			# 	self.__tempdict[key] = list()

			comm = self.__get_commision_on_services(data.service,data.service_price,
											 		data.transaction_type,data.add_ons_price)

			total += comm
			total2 += data.total_amount

			data_dict = dict()
			data_dict = {
				'datecreated': key,
				'clientname': data.client_name,
				'services': data.service + ', ' + data.add_ons,
				'timespent': '{:02d}H:{:02d}M'.format(*divmod(data.time_spent, 60)),
				'amountpaid': data.total_amount,
				'commision_on_service': comm
			}

			self.__tempdict['data'].append(data_dict)

		return total, total2

	def __get_commision_on_services(self, service, serviceprice, trantype, add_ons_price):
		commlist = ['Hot Stone Massage', 'Shiatsu Massage']
		retval = 0

		servicecomm = 0
		if service in commlist:
			if serviceprice is None:
				serviceprice = 0

			servicecomm = int(serviceprice) * 0.10

		if not add_ons_price:
			return servicecomm

		addonsprices = add_ons_price.split(',')
		addonscomm = 0

		if trantype in ['Walk-In','Non-Member']:		
			for price in addonsprices:
				if int(price) <= 299:
					addonscomm += 25
				elif int(price) >= 300:
					addonscomm += 50

		elif trantype == 'Member':
			for price in addonsprices:
				if int(price) <= 149:
					addonscomm += 25
				elif int(price) >= 150:
					addonscomm += 50

		retval = addonscomm + servicecomm

		return retval

	def __get_members(self, ds, de, attid):
		retval = list()
		search_filter = list()

		search_filter.append(getattr(T_Member00, 'datecreated').between(ds, de))
		search_filter.append(getattr(T_Member00, 'attendantid') == attid)

		result = self.__session.query(T_Member00).filter(and_(*search_filter)).all()

		total = 0
		total2 = 0

		for data in result:
			key = data.datecreated.strftime('%B-%d-%Y')

			# if key not in self.__tempdict:
			# 	self.__tempdict[key] = list()

			if data.upgraded:
				total += 25
				total2 += 300

				data_dict = dict()
				data_dict = {
					'clientname': data.name,
					'services': 'Membership Sold - ' + 'Personalized',
					'amountpaid': 300,
					'commision_on_service': 25,
					'timespent': 'N/A',
					'datecreated':key
				}
			else:

				comm = 25 if data.membertype == 'Personalized' else 50

				total += comm
				total2 += data.membershipcost

				data_dict = dict()
				data_dict = {
					'clientname': data.name,
					'services': 'Membership Sold - ' + data.membertype,
					'amountpaid': data.membershipcost,
					'commision_on_service': comm,
					'timespent': 'N/A',
					'datecreated':key
				}

			self.__tempdict['data'].append(data_dict)

		return total, total2

	def __get_members_upg(self, ds, de, attid):
		retval = list()
		search_filter = list()

		search_filter.append(getattr(T_Member00, 'upgraded').between(ds, de))
		search_filter.append(getattr(T_Member00, 'upgraded_by') == attid)

		result = self.__session.query(T_Member00).filter(and_(*search_filter)).all()

		total = 0
		total2 = 0

		for data in result:
			key = data.upgraded.strftime('%B-%d-%Y')

			# if key not in self.__tempdict:
			# 	self.__tempdict[key] = list()

			total += 25
			total2 += 300

			data_dict = dict()
			data_dict = {
				'clientname': data.name,
				'services': 'Membership Sold - ' + data.membertype + '[UPGRADED]',
				'amountpaid': 300,
				'commision_on_service': 25,
				'timespent': 'N/A',
				'datecreated':key
			}

			self.__tempdict['data'].append(data_dict)

		return total, total2


class Summary_report(object):
	def __init__(self, **kwargs):
		self.__session = Session()
		self.__args = kwargs
		self.__retval = dict()

		# column start for insert of image
		self.__column = 0

	def get_reports_data(self):
		self.__get_attendants()
		self.__insert_dates()
		self.__get_attendants01()
		self.__get_transactions()
		self.__get_membership()
		self.__get_upgraded()

		total_dict = self.__make_total_dict(self.__retval)

		filename = self.__generate_excelfil(self.__retval, total_dict)

		return filename

	def __get_attendants(self):
		result = self.__session.query(T_Attendants).all()

		for d in result:
			self.__retval[d.attendantid] = {
				'attendant_name': d.attendant_name,
				'allowance_perday': d.allowance,
				'allowance_total': 0,
				'addons_total': 0,
				'commision_total_addons': 0,
				'services_total': 0,
				'commision_total_service': 0,
				'personalized_total': 0,
				'incentive_total_personalized': 0,
				'family_total': 0,
				'incentive_total_family': 0,
				'upgraded_total': 0,
				'incentive_total_upgraded': 0,
				'dates': dict()
			}

	def __insert_dates(self):
		dateend = datetime.datetime.strptime(self.__args['to'], '%Y-%m-%d')
		delta = datetime.timedelta(days=1)

		for attendantid in self.__retval.keys():
			datestart = datetime.datetime.strptime(self.__args['from'], '%Y-%m-%d')

			while datestart <= dateend:
				self.__retval[attendantid]['dates'][datestart.strftime('%B-%d-%Y')] = 'AWOL'

				datestart += delta


	def __get_attendants01(self):
		search_filter = [getattr(T_Attendants01, 'trandate').between(self.__args['from'], self.__args['to'])]

		result = self.__session.query(T_Attendants01).filter(*search_filter).order_by(
			T_Attendants01.attendantid.asc(), T_Attendants01.trandate.asc()).all()

		for d in result:
			if not d.timein:
				timein = 'No Time In'
			else:
				timein = d.timein

			if not d.timeout:
				timeout = 'No Time Out'
			else:
				timeout = d.timeout

			self.__retval[d.attendantid]['dates'][d.trandate.strftime('%B-%d-%Y')] = timein + ' - ' + timeout

			x = self.__compute_allowance(d.timein, d.timeout, self.__retval[d.attendantid]['allowance_perday'])
			self.__retval[d.attendantid]['allowance_total'] += x

	def __compute_allowance(self, timein, timeout, allowance):

		if any([not timein, not timeout]):
			return 0

		per_hour = allowance / 12

		timein = datetime.datetime.strptime(timein, '%I:%M %p')
		timeout = datetime.datetime.strptime(timeout, '%I:%M %p')
		timediff = timeout - timein
		hours = int(timediff.seconds / 3600)
		retval = int(per_hour * hours)

		return retval

	def __get_transactions(self):
		search_filter = [getattr(T_Transaction, 'datecreated').between(self.__args['from'], self.__args['to'])]

		result = self.__session.query(T_Transaction).filter(*search_filter)

		temp_dict = dict()
		for d in result:
			comms = self.__compute_commision(d.service,d.service_price,d.transaction_type,d.add_ons_price)
			self.__retval[d.attendantid]['commision_total_service'] += comms[0]
			self.__retval[d.attendantid]['commision_total_addons'] += comms[1]
			self.__retval[d.attendantid]['addons_total'] += self.__compute_addons(d.add_ons_price)
			self.__retval[d.attendantid]['services_total'] += self.__compute_services(d.service_price)

	def __compute_services(self, serviceprice):
		return int(serviceprice)

	def __compute_addons(self, addonsprice):
		total = 0

		if not addonsprice:
			return 0

		for price in addonsprice.split(','):
			total += int(price)

		return total

	def __compute_commision(self, service, serviceprice, trantype, add_ons_price):
		ignore_list = ['5-in-1 Signature Massage', 'Relaxing Swedish Massage']
		retval = list()

		# get services commisions
		servicecomm = 0
		if service not in ignore_list:
			if serviceprice is None:
				serviceprice = 0

			servicecomm = int(serviceprice) * 0.10

		retval.append(servicecomm)

		# get addons commisions
		addonscomm = 0
		if not add_ons_price:
			addonscomm = 0
		else:
			addonsprices = add_ons_price.split(',')
			if trantype in ['Walk-In','Non-Member']:		
				for price in addonsprices:
					if int(price) <= 299:
						addonscomm += 25
					elif int(price) >= 300:
						addonscomm += 50

			elif trantype == 'Member':
				for price in addonsprices:
					if int(price) <= 149:
						addonscomm += 25
					elif int(price) >= 150:
						addonscomm += 50

		retval.append(addonscomm)

		return retval

	def __get_membership(self):
		search_filter = [getattr(T_Member00, 'datecreated').between(self.__args['from'], self.__args['to'])]
		search_filter.append(getattr(T_Member00, 'attendantid') != 0)

		result = self.__session.query(T_Member00).filter(*search_filter).all()

		for d in result:
			if any([d.upgraded, d.upgraded_by]):
				self.__retval[d.attendantid]['personalized_total'] += d.membershipcost / 2
				self.__retval[d.attendantid]['incentive_total_personalized'] += 25
				continue

			if d.membertype == 'Personalized':
				self.__retval[d.attendantid]['personalized_total'] += d.membershipcost
				self.__retval[d.attendantid]['incentive_total_personalized'] += 25
			elif d.membertype == 'Family':
				self.__retval[d.attendantid]['family_total'] += d.membershipcost
				self.__retval[d.attendantid]['incentive_total_family'] += 50

	def __get_upgraded(self):
		search_filter = [getattr(T_Member00, 'upgraded').between(self.__args['from'], self.__args['to'])]		

		result = self.__session.query(T_Member00).filter(*search_filter).all()

		for d in result:
			self.__retval[d.upgraded_by]['upgraded_total'] += d.membershipcost / 2
			self.__retval[d.upgraded_by]['incentive_total_upgraded'] += 25

	def __make_total_dict(self, data):
		retval = {
			'total_allowance': 0,
			'total_commision': 0,
			'total_incentive': 0,
			'total_gross_sales': 0,
			'operational_expenses': 0,
			'total_net_sales': 0,
			'total_addons': 0,
			'total_services': 0,
			'total_members_p': 0,
			'total_members_f': 0,
			'total_members_u': 0
		}

		for key, value in data.items():
			retval['total_allowance'] += value.get('allowance_total', 0)
			retval['total_commision'] += sum([value.get('commision_total_addons', 0), value.get('commision_total_service', 0)])
			retval['total_incentive'] += sum([value.get('incentive_total_personalized', 0),
											  value.get('incentive_total_family', 0),
											  value.get('incentive_total_upgraded', 0)])
			retval['total_gross_sales'] += sum([value.get('addons_total', 0),
											    value.get('services_total', 0),
											    value.get('personalized_total', 0),
											    value.get('family_total', 0),
											    value.get('upgraded_total', 0)])
			retval['total_addons'] +=  value.get('addons_total', 0)
			retval['total_services'] +=  value.get('services_total', 0)
			retval['total_members_p'] +=  value.get('personalized_total', 0)
			retval['total_members_f'] +=  value.get('family_total', 0)
			retval['total_members_u'] +=  value.get('upgraded_total', 0)

		retval['total_net_sales'] = (retval['total_gross_sales'] - 
									 retval['total_incentive'] - 
									 retval['total_commision'] - 
									 retval['total_allowance'] -
									 retval['operational_expenses'])

		return retval

	def __generate_excelfil(self, attendantdata, totaldata):
		wb = Workbook()
		ws = wb.active

		ws.title = "Summary Report"
		ws.sheet_properties.tabColor = '1072BA'

		self.__insert_first_column(ws, attendantdata)
		self.__insert_all_data_to_excel(ws, attendantdata)
		self.__insert_totals(ws, totaldata)

		filename = datetime.datetime.now().strftime('%Y%m%d%H%M%S')

		folderpath = os.path.join(os.getcwd(), 'Reports')
		if not os.path.exists(folderpath):
			os.makedirs(folderpath)

		retval = 'Sales_report_' + filename + '.xlsx'
		wb.save(os.path.join(folderpath, retval))

		return retval

	def __insert_first_column(self, ws, data):
		# insert dates to first column
		for key, value in data.items():
			row_count = 2

			dates = list(value['dates'].keys())
			dates.sort(key=lambda x: datetime.datetime.strptime(x, '%B-%d-%Y'))

			for date in dates:
				cell = ws.cell(row=row_count, column=1)
				cell.value = date

				# set fonts
				ft = Font(bold=True)
				al = Alignment(horizontal='center', vertical='center')
				cell.font = ft
				cell.alignment = al

				# set width
				coordinate = cell.coordinate
				coor = re.split('(\d+)', coordinate)
				ws.column_dimensions[str(coor[0])].width = 25

				row_count += 1

			additionalheader = ['Allowance', 'Commision on Service', 'Incentive on Membership', 'Total per Attendant']

			row_count += 1
			for header in additionalheader:
				cell = ws.cell(row=row_count, column=1)
				cell.value = header

				# set fonts
				ft = Font(bold=True)
				al = Alignment(horizontal='center', vertical='center')
				cell.font = ft
				cell.alignment = al

				row_count += 1

			break

	def __insert_all_data_to_excel(self, ws, data):
		
		column_count = 2
		for key, value in data.items():
			row_count = 1

			dates = list(value['dates'].keys())
			dates.sort(key=lambda x: datetime.datetime.strptime(x, '%B-%d-%Y'))
			dates.insert(0, 'attendant_name')

			for date in dates:
				cell = ws.cell(row=row_count, column=column_count)

				if date == 'attendant_name':
					cell.value = value[date]

					# set fonts
					ft = Font(bold=True)
					al = Alignment(horizontal='center', vertical='center')
					cell.font = ft
					cell.alignment = al

					coordinate = cell.coordinate
					coor = re.split('(\d+)', coordinate)
					# set height and width of headers cell
					ws.row_dimensions[int(coor[1])].height = 45
					ws.column_dimensions[str(coor[0])].width = 25

				else:
					# set font
					al = Alignment(horizontal='center', vertical='center')
					cell.alignment = al

					cell.value = value['dates'][date]

				row_count += 1

			additionalheader = ['Allowance', 'Commision on Service', 'Incentive on Membership', 'Total per Attendant']
			row_count += 1
			for header in additionalheader:
				cell = ws.cell(row=row_count, column=column_count)
				# set font
				al = Alignment(horizontal='center', vertical='center')
				cell.alignment = al

				if header == 'Allowance':
					cell.value = value['allowance_total']

				elif header == 'Commision on Service':
					val = sum([value.get('commision_total_addons', 0), value.get('commision_total_service', 0)])
					cell.value = val

				elif header == 'Incentive on Membership':
					val = sum([value.get('incentive_total_personalized', 0),
							   value.get('incentive_total_family', 0),
							   value.get('incentive_total_upgraded', 0)])
					cell.value = val

				elif header == 'Total per Attendant':
					val = sum([value['allowance_total'],
							   value.get('commision_total_addons', 0),
							   value.get('commision_total_service', 0),
							   value.get('incentive_total_personalized', 0),
							   value.get('incentive_total_family', 0),
							   value.get('incentive_total_upgraded', 0)])
					cell.value = val

				row_count += 1

			column_count += 1

		self.__column = column_count + 1

	def __insert_totals(self, ws, totaldata):
		# insert banahaw image on start
		cell = ws.cell(row=1, column=1)
		img_path = os.path.join(os.getcwd(), 'BanahawApp', 'static', 'img', 'Untitled2.png')
		img = drawing.image.Image(img_path)
		ws.add_image(img, cell.coordinate)

		# insert banahaw image
		cell = ws.cell(row=1, column=self.__column)
		img_path = os.path.join(os.getcwd(), 'BanahawApp', 'static', 'img', 'Untitled.png')
		img = drawing.image.Image(img_path)
		ws.add_image(img, cell.coordinate)

		# insert color under the image
		cell1 = ws.cell(row=6, column=self.__column)
		cell2 = ws.cell(row=6, column=self.__column + 5)
		mergecells = '{}:{}'.format(cell1.coordinate, cell2.coordinate)
		ws.merge_cells(mergecells)
		fl = PatternFill(start_color='2E8B57', end_color='2E8B57', fill_type='solid')
		cell1.fill = fl

		# insert totals
		row_count = 7

		# add blank for blank in excel
		loop = list(totaldata.keys())
		loop.append('blank')
		loop.append('blank')

		for index, key in enumerate(loop):

			if index == 0:			
				self.__merge_cell_insert_head(ws, row_count, 'Total On New Members Personalized', '993366')
				self.__merge_cell_insert_data(ws, row_count, totaldata['total_members_p'], '993366')
			elif index == 1:
				self.__merge_cell_insert_head(ws, row_count, 'Total On New Members Family', '994d33')
				self.__merge_cell_insert_data(ws, row_count, totaldata['total_members_f'], '994d33')
			elif index == 2:
				self.__merge_cell_insert_head(ws, row_count, 'Total On Upgraded Members to Family', '996633')
				self.__merge_cell_insert_data(ws, row_count, totaldata['total_members_u'], '996633')
			elif index == 3:
				self.__merge_cell_insert_head(ws, row_count, 'Total On Services/Packages/Promos', '998033')
				self.__merge_cell_insert_data(ws, row_count, totaldata['total_services'], '998033')
			elif index == 4:
				self.__merge_cell_insert_head(ws, row_count, 'Total On Add - Ons', '999933')
				self.__merge_cell_insert_data(ws, row_count, totaldata['total_addons'], '999933')
			elif index == 5:
				# blank
				cell1 = ws.cell(row=row_count, column=self.__column)
				cell2 = ws.cell(row=row_count, column=self.__column + 5)
				mergecells = '{}:{}'.format(cell1.coordinate, cell2.coordinate)
				ws.merge_cells(mergecells)
			elif index == 6:
				self.__merge_cell_insert_head(ws, row_count, 'Total Allowance', '339980')
				self.__merge_cell_insert_data(ws, row_count, totaldata['total_allowance'], '339980')
			elif index == 7:
				self.__merge_cell_insert_head(ws, row_count, 'Total Commision', '339999')
				self.__merge_cell_insert_data(ws, row_count, totaldata['total_commision'], '339999')
			elif index == 8:
				self.__merge_cell_insert_head(ws, row_count, 'Total Incentive on Membership', '334d99')
				self.__merge_cell_insert_data(ws, row_count, totaldata['total_incentive'], '334d99')
			elif index == 9:
				self.__merge_cell_insert_head(ws, row_count, 'Operational Expenses', '4d3399')
				self.__merge_cell_insert_data(ws, row_count, totaldata['operational_expenses'], '4d3399')
			elif index == 10:
				# blank
				cell1 = ws.cell(row=row_count, column=self.__column)
				cell2 = ws.cell(row=row_count, column=self.__column + 5)
				mergecells = '{}:{}'.format(cell1.coordinate, cell2.coordinate)
				ws.merge_cells(mergecells)
			elif index == 11:
				self.__merge_cell_insert_head(ws, row_count, 'Total Gross Sales', '99334d')
				self.__merge_cell_insert_data(ws, row_count, totaldata['total_gross_sales'], '99334d')
			elif index == 12:
				self.__merge_cell_insert_head(ws, row_count, 'Total Net Sales', '339966')
				self.__merge_cell_insert_data(ws, row_count, totaldata['total_net_sales'], '339966')

			row_count += 1

		else:
			cell1 = ws.cell(row=row_count, column=self.__column)
			cell2 = ws.cell(row=row_count, column=self.__column + 5)
			mergecells = '{}:{}'.format(cell1.coordinate, cell2.coordinate)
			ws.merge_cells(mergecells)
			fl = PatternFill(start_color='2E8B57', end_color='2E8B57', fill_type='solid')
			cell1.fill = fl



	def __merge_cell_insert_head(self, ws, row, value, color):
		cell1 = ws.cell(row=row, column=self.__column)
		cell2 = ws.cell(row=row, column=self.__column + 3)
		mergecells = '{}:{}'.format(cell1.coordinate, cell2.coordinate)
		ws.merge_cells(mergecells)
		ft = Font(bold=True, italic=True, color=color)
		al = Alignment(horizontal='center', vertical='center')
		cell1.font = ft
		cell1.alignment = al
		cell1.value = value

	def __merge_cell_insert_data(self, ws, row, value, color):
		cell1 = ws.cell(row=row, column=self.__column + 4)
		cell2 = ws.cell(row=row, column=self.__column + 5)
		mergecells = '{}:{}'.format(cell1.coordinate, cell2.coordinate)
		ws.merge_cells(mergecells)
		ft = Font(bold=True, italic=True, color=color)
		al = Alignment(horizontal='center', vertical='center')
		cell1.font = ft
		cell1.alignment = al
		cell1.value = value


class Memberslist_report(object):
	def __init__(self, **kwargs):
		self.__session = Session()
		self.__args = kwargs
		self.__retval = dict()

	def get_reports_data(self):
		datestart = self.__args['from']
		dateend = self.__args['to']

		memberslist = self.__get_members(datestart, dateend)
		print(memberslist)
		retval = self.__generate_excelfile(memberslist)

		return retval

	def __generate_excelfile(self, memberslist):
		wb = Workbook()
		ws = wb.active

		ws.title = "Personalized"
		wb.create_sheet(index=1, title='Family')
		wb.create_sheet(index=2, title='Upgraded')

		personalizedsheet = wb.get_sheet_by_name('Personalized')
		familysheet = wb.get_sheet_by_name('Family')
		upgradedsheet = wb.get_sheet_by_name('Upgraded')

		personalizeddata = memberslist.get('Personalized', list())
		familydata = memberslist.get('Family', list())
		upgradeddata = memberslist.get('Upgraded', list())

		self.__extract_personalized_data(personalizedsheet, personalizeddata)
		self.__extract_family_data(familysheet, familydata)
		self.__extract_upgraded_data(upgradedsheet, upgradeddata)

		# make file
		filename = datetime.datetime.now().strftime('%Y%m%d%H%M%S')

		folderpath = os.path.join(os.getcwd(), 'Reports')
		if not os.path.exists(folderpath):
			os.makedirs(folderpath)

		retval = 'Members_list' + filename + '.xlsx'
		wb.save(os.path.join(folderpath, retval))

		return retval

	def __extract_personalized_data(self, sheet, data):
		headers = ['name', 'address', 'birthdate',
				   'mobile_number', 'landline_number',
				   'email_address', 'date_applied']

		ft = Font(bold=True)
		row_count = 1
		for column_count, header in enumerate(headers):
			cell = sheet.cell(row=row_count, column=column_count + 1)
			cell.value = header.upper()
			cell.font = ft

		row_count2 = 2
		for data_dict in data:

			for column_count, header in enumerate(headers):
				cell = sheet.cell(row=row_count2, column=column_count + 1)
				cell.value = data_dict[header]

			row_count2 += 1


	def __extract_family_data(self, sheet, data):
		headers = ['name', 'address', 'birthdate',
				   'mobile_number', 'landline_number',
				   'email_address', 'submembers', 'date_applied']

		ft = Font(bold=True)
		row_count = 1
		for column_count, header in enumerate(headers):
			cell = sheet.cell(row=row_count, column=column_count + 1)
			cell.value = header.upper()
			cell.font = ft

		row_count2 = 2
		for data_dict in data:

			for column_count, header in enumerate(headers):
				cell = sheet.cell(row=row_count2, column=column_count + 1)
				cell.value = data_dict[header]

			row_count2 += 1

	def __extract_upgraded_data(self, sheet, data):
		headers = ['name', 'address', 'birthdate',
				   'mobile_number', 'landline_number',
				   'email_address', 'submembers', 'date_upgraded']

		ft = Font(bold=True)
		row_count = 1
		for column_count, header in enumerate(headers):
			cell = sheet.cell(row=row_count, column=column_count + 1)
			cell.value = header.upper()
			cell.font = ft

		row_count2 = 2
		for data_dict in data:

			for column_count, header in enumerate(headers):
				cell = sheet.cell(row=row_count2, column=column_count + 1)
				cell.value = data_dict[header]

			row_count2 += 1

	def __get_members(self, ds, de):
		retval = dict()
		search_filter = [getattr(T_Member00, 'datecreated').between(self.__args['from'], self.__args['to'])]
		search_filter.append(getattr(T_Member00, 'attendantid') != 0)

		result = self.__session.query(T_Member00).filter(*search_filter).all()

		for data in result:
			temp_dict = dict()

			if data.upgraded or data.upgraded_by:
				if 'Upgraded' not in retval:
					retval['Upgraded'] = list()

				temp_dict['name'] = data.name
				temp_dict['address'] = data.address
				temp_dict['mobile_number'] = data.mobile_number
				temp_dict['landline_number'] = data.landline_number
				temp_dict['email_address'] = data.email_address
				temp_dict['birthdate'] = data.birthdate
				temp_dict['date_applied'] = data.datecreated
				temp_dict['date_upgraded'] = data.upgraded

				member01list = list()
				member01res = self.__session.query(T_Member01).filter(T_Member01.member00id == data.member00id).all()
				for data2 in member01res:
					member01list.append(data2.name)

				temp_dict['submembers'] = ','.join(member01list)

				retval['Upgraded'].append(temp_dict)

			else:

				if data.membertype not in retval:
					retval[data.membertype] = list()

				temp_dict['name'] = data.name
				temp_dict['address'] = data.address
				temp_dict['mobile_number'] = data.mobile_number
				temp_dict['landline_number'] = data.landline_number
				temp_dict['email_address'] = data.email_address
				temp_dict['birthdate'] = data.birthdate
				temp_dict['date_applied'] = data.datecreated

				if data.membertype == 'Family':
					member01list = list()
					member01res = self.__session.query(T_Member01).filter(T_Member01.member00id == data.member00id).all()
					for data2 in member01res:
						member01list.append(data2.name)

					temp_dict['submembers'] = ','.join(member01list)					

				retval[data.membertype].append(temp_dict)

		return retval
